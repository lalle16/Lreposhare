from __future__ import annotations

import io
import os
from typing import Tuple

import openpyxl
import pandas as pd

from .config import get_validations_excel_path, client_data_path


def _get_bytes(file_or_path) -> bytes | None:
    """Return bytes for a file-like object without exhausting the stream; else None.
    Prefers getvalue() (Streamlit UploadedFile), falls back to seek(0)+read().
    """
    if hasattr(file_or_path, "getvalue"):
        try:
            return file_or_path.getvalue()
        except Exception:
            pass
    if hasattr(file_or_path, "read"):
        try:
            # Reset pointer if possible to avoid empty reads across iterations
            if hasattr(file_or_path, "seek"):
                try:
                    file_or_path.seek(0)
                except Exception:
                    pass
            return file_or_path.read()
        except Exception:
            return None
    return None


def load_excel_data(file_or_path, sheet_name: str | None = None, table_name: str | None = None) -> pd.DataFrame:
    """
    Load an Excel file, optionally from a specific sheet or named table.
    - If table_name is None, load the sheet from row 0 (header row).
    - If table_name is provided, load only the named table from the sheet.
    Accepts a file-like object (as provided by Streamlit) or a path.
    """

    data_bytes = _get_bytes(file_or_path)

    def _resolve_sheet_title(wb: openpyxl.workbook.workbook.Workbook, name: str | None) -> str:
        if not name:
            return wb.active.title
        target = str(name).strip().casefold()
        for title in wb.sheetnames:
            if str(title).strip().casefold() == target:
                return title
        raise ValueError(f"SHEET_NOT_FOUND: Sheet '{name}' not found. Available: {', '.join(wb.sheetnames)}")

    if table_name is not None and str(table_name).strip() != "" and not pd.isna(table_name):
        # Use openpyxl to find the named table range
        if data_bytes is not None:
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True)
        else:
            wb = openpyxl.load_workbook(file_or_path, data_only=True)

        actual_title = _resolve_sheet_title(wb, sheet_name)
        ws = wb[actual_title]
        if table_name not in ws.tables:
            raise ValueError(f"TABLE_NOT_FOUND: Table '{table_name}' not found in sheet '{ws.title}'")
        table = ws.tables[table_name]
        ref = table.ref  # e.g., 'B2:F10'
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
        data = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        data = list(data)
        if not data:
            return pd.DataFrame()
        header = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=header)
        return df

    # No table name: let pandas read the sheet
    if data_bytes is not None:
        if sheet_name is None:
            df = pd.read_excel(io.BytesIO(data_bytes), header=0)
        else:
            # Resolve to actual sheet title to ignore case/whitespace mismatches
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True)
            actual_title = _resolve_sheet_title(wb, sheet_name)
            df = pd.read_excel(io.BytesIO(data_bytes), sheet_name=actual_title, header=0)
    else:
        if sheet_name is None:
            df = pd.read_excel(file_or_path, header=0)
        else:
            # Resolve using workbook loaded from path
            wb = openpyxl.load_workbook(file_or_path, data_only=True)
            actual_title = _resolve_sheet_title(wb, sheet_name)
            df = pd.read_excel(file_or_path, sheet_name=actual_title, header=0)
    return df

def load_backend_validations() -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load validation rules and error messages from the backend validations.xlsx.
    These are not user-editable in the app.
    Returns (validation_sheet, error_messages)
    """
    validations_path = get_validations_excel_path(client_data_path)
    validation_sheet = pd.read_excel(validations_path, sheet_name="Columns", header=0)
    error_messages = pd.read_excel(validations_path, sheet_name="Messages", header=0)
    return validation_sheet, error_messages
